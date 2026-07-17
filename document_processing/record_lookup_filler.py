#!/usr/bin/env python3
"""
record_lookup_filler.py

Fills a blank "Value" column in a target Excel workbook by matching each
row's "ID" against a source CSV export and pulling the corresponding value
from that CSV.

Why this exists
----------------
Many workflows involve two disconnected files: an internal tracking
workbook (.xlsx) with a lookup column that starts out blank, and a raw
system export (.csv) that has the answer somewhere in it — but with no
header row, dozens of columns, and some rows still "in progress" with a
placeholder instead of a real value. Reconciling these by hand with
VLOOKUP is slow and error-prone. This script automates the lookup and
prints a clear summary of what happened (matched / already filled /
unmatched).

How matching works
-------------------
1. The source CSV is read with no header row assumed. Column A (index 0)
   is treated as the ID, and Column N (index 13) as the Value.
2. Rows where the Value is a placeholder (e.g. "NA", meaning "not yet
   available") are skipped, so an in-progress row never gets a bogus value.
3. The target workbook is read by *header name*, not fixed column letters,
   so it keeps working even if columns are reordered or new ones are added.
4. Only blank cells in the Value column are written to; anything already
   filled is left untouched and reported separately.

Usage
-----
    python record_lookup_filler.py --excel tracker.xlsx --csv export.csv

Optional output path (defaults to "<excel>_updated.xlsx"):

    python record_lookup_filler.py \\
        --excel tracker.xlsx \\
        --csv export.csv \\
        --output tracker_filled.xlsx

Requirements
------------
    pip install -r requirements.txt   # installs openpyxl

Column index assumptions (CSV_ID_COL / CSV_VALUE_COL below) are based on a
specific export layout. If your export uses different columns, adjust
those two constants. Likewise, the header names the script looks for in
the target workbook (ID_HEADER / VALUE_HEADER) can be changed to match
your own file.
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
from dataclasses import dataclass

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# --- Column positions in the source CSV (0-indexed, no header row) ---------
CSV_ID_COL = 0       # Column A: unique identifier
CSV_VALUE_COL = 13   # Column N: value to pull across

# --- Column header names expected in the target workbook -------------------
ID_HEADER = "ID"
VALUE_HEADER = "Value"

# --- Placeholder value in the CSV meaning "not yet available" --------------
PENDING_PLACEHOLDER = "NA"

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)


@dataclass
class FillSummary:
    """Result counts from a fill run, useful for logging or testing."""
    total_mappings_loaded: int
    matched: int
    already_filled: int
    unmatched: int


def build_lookup(csv_path: str) -> dict[str, str]:
    """Build an {ID: Value} lookup from the source CSV.

    Rows with a missing ID, a missing Value, or a pending placeholder
    ("NA") in the Value field are skipped.
    """
    lookup: dict[str, str] = {}
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if len(row) <= max(CSV_ID_COL, CSV_VALUE_COL):
                continue
            record_id = (row[CSV_ID_COL] or "").strip()
            value = (row[CSV_VALUE_COL] or "").strip()
            if record_id and value and value.upper() != PENDING_PLACEHOLDER:
                lookup[record_id] = value
    return lookup


def _find_header_columns(ws: Worksheet) -> tuple[int, int]:
    """Locate the ID / Value columns by header name in row 1."""
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    try:
        return headers[ID_HEADER], headers[VALUE_HEADER]
    except KeyError as exc:
        raise ValueError(
            f"Could not find required column {exc}. "
            f"Headers found in row 1: {list(headers.keys())}"
        ) from exc


def fill_values(excel_path: str, csv_path: str, output_path: str) -> FillSummary:
    """Read `excel_path`, fill blank Value cells using `csv_path`, save to `output_path`."""
    lookup = build_lookup(csv_path)
    log.info("Loaded %d ID -> Value mappings from CSV.", len(lookup))

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    id_col, value_col = _find_header_columns(ws)

    matched = unmatched = already_filled = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        id_cell = row[id_col - 1]
        value_cell = row[value_col - 1]

        record_id = str(id_cell.value).strip() if id_cell.value else ""
        if not record_id:
            continue

        if value_cell.value not in (None, ""):
            already_filled += 1
            continue

        value = lookup.get(record_id)
        if value:
            value_cell.value = value
            matched += 1
        else:
            unmatched += 1

    wb.save(output_path)

    summary = FillSummary(
        total_mappings_loaded=len(lookup),
        matched=matched,
        already_filled=already_filled,
        unmatched=unmatched,
    )

    log.info("Matched & filled: %d", summary.matched)
    log.info("Already had a value (left untouched): %d", summary.already_filled)
    log.info("No match found in CSV: %d", summary.unmatched)
    log.info("Saved to %s", output_path)

    return summary


def _default_output_path(excel_path: str) -> str:
    base, ext = os.path.splitext(excel_path)
    return f"{base}_updated{ext}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill a blank 'Value' column in a workbook from a source CSV.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--excel", required=True, help="Path to the target .xlsx workbook")
    parser.add_argument("--csv", required=True, help="Path to the source .csv export")
    parser.add_argument(
        "--output",
        default=None,
        help="Path for the updated workbook (default: <excel>_updated.xlsx)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if not os.path.isfile(args.excel):
        raise FileNotFoundError(f"Excel file not found: {args.excel}")
    if not os.path.isfile(args.csv):
        raise FileNotFoundError(f"CSV file not found: {args.csv}")

    output_path = args.output or _default_output_path(args.excel)
    fill_values(args.excel, args.csv, output_path)


if __name__ == "__main__":
    main()
