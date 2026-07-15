#!/usr/bin/env python3
"""
Printed Report — PDF to Excel Converter
Usage: python pdf_to_excel_extractor.py <input.pdf> [output.xlsx]
"""

import sys
import re
import os
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("Installing pdfplumber...")
    os.system(f"{sys.executable} -m pip install pdfplumber --break-system-packages -q")
    import pdfplumber

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Formatting helpers ──────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="D6E4F0")   # light blue
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT     = Font(name="Arial", size=10)
TITLE_FONT    = Font(name="Arial", bold=True, size=13)
SUB_FONT      = Font(name="Arial", size=10, italic=True)

THIN = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_HEADERS = ["Sr.No.", "File No.", "Applicant Name",
               "Document No.", "Birth Date", "Reference No.", "Printed By"]
COL_WIDTHS  = [8, 18, 32, 16, 14, 14, 14]


def extract_metadata(pdf_path: str) -> dict:
    """Pull report title, generation date, and total-records line from page text."""
    meta = {"title": "", "org_unit": "", "report_type": "",
            "generated_on": "", "total_records": ""}
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if lines:
        meta["org_unit"] = lines[0]            # e.g. org/branch identifier line
    if len(lines) > 1:
        meta["report_type"] = lines[1]         # e.g. "Printed"
    for line in lines:
        if line.startswith("Report Generated on"):
            meta["generated_on"] = line.split(":", 1)[-1].strip()
        if line.startswith("Total Records"):
            meta["total_records"] = line.split(":", 1)[-1].strip()
    meta["title"] = f"{meta['org_unit']} — {meta['report_type']}"
    return meta


def extract_table_rows(pdf_path: str) -> list[list[str]]:
    """Return all data rows (no header) from every page, newlines collapsed."""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for i, row in enumerate(table):
                    if i == 0:          # skip header row on each page
                        continue
                    cleaned = [
                        (cell or "").replace("\n", " ").strip()
                        for cell in row
                    ]
                    if any(cleaned):
                        rows.append(cleaned)
    return rows


def build_excel(rows: list[list[str]], meta: dict, out_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Printed Report"

    # ── Title block ────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = meta["title"]
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G2")
    ws["A2"] = (f"Report Generated on: {meta['generated_on']}    |    "
                f"Total Records: {meta['total_records']}")
    ws["A2"].font = SUB_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # blank separator
    ws.row_dimensions[3].height = 6

    # ── Column headers (row 4) ─────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(zip(COL_HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 20

    # ── Data rows ──────────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows, start=5):
        fill = ALT_ROW_FILL if r_idx % 2 == 0 else WHITE_FILL
        for c_idx, value in enumerate(row[:7], start=1):
            # keep Sr.No. as integer when possible
            if c_idx == 1:
                try:
                    value = int(value)
                except ValueError:
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if c_idx in (1, 4, 5, 6, 7) else "left",
                vertical="center"
            )
        ws.row_dimensions[r_idx].height = 15

    # ── Freeze panes below header ──────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Footer ────────────────────────────────────────────────────────────
    footer_row = 5 + len(rows)
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    ws[f"A{footer_row}"] = (f"Generated by: {meta.get('generated_by', 'SYSTEM')}    |    "
                            f"Report Date: {meta['generated_on']}")
    ws[f"A{footer_row}"].font = Font(name="Arial", size=9, italic=True, color="666666")
    ws[f"A{footer_row}"].alignment = Alignment(horizontal="right")

    wb.save(out_path)
    print(f"✅  Saved: {out_path}  ({len(rows)} records)")


def convert(pdf_path: str, out_path: str | None = None) -> str:
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    if out_path is None:
        out_path = str(Path(pdf_path).with_suffix(".xlsx"))

    print(f"📄  Reading: {pdf_path}")
    meta = extract_metadata(pdf_path)
    rows = extract_table_rows(pdf_path)

    # Try to get "Printed By" username for footer from last line of PDF text
    with pdfplumber.open(pdf_path) as pdf:
        last_text = pdf.pages[-1].extract_text() or ""
    m = re.search(r"^([A-Z]+)\s+-\s+Generated On", last_text, re.MULTILINE)
    if m:
        meta["generated_by"] = m.group(1)

    print(f"📊  Found {len(rows)} data rows")
    build_excel(rows, meta, out_path)
    return out_path


# ── CLI entry-point ────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python pdf_to_excel_extractor.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_in  = sys.argv[1]
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else None
    convert(pdf_in, xlsx_out)
