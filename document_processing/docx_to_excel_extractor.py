"""
Application Form DOCX → Excel Extractor
==========================================
Parses a structured Application Form Word document (.docx) and extracts
applicant data into a structured Excel spreadsheet.

Usage:
    python docx_to_excel_extractor.py <input.docx> [output.xlsx]

If output path is omitted, the script writes to the same folder
as the input file with the suffix _extracted.xlsx.
"""

import sys
import re
from pathlib import Path
from datetime import datetime

from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(value: str) -> str:
    """Strip whitespace and asterisks (markdown bold artefacts)."""
    return value.strip().strip("*").strip()


def strip_phone(value: str) -> str:
    """Remove leading + and non-digit characters for a plain number."""
    if not value:
        return value
    # Keep only digits
    digits = re.sub(r"\D", "", value)
    return digits if digits else value


def parse_date(value: str) -> str:
    """
    Try to convert various date formats to DD-Mon-YYYY (e.g. 01-Feb-2002).
    Falls back to the raw string if parsing fails.
    """
    if not value:
        return value
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%b/%Y"):
        try:
            dt = datetime.strptime(value.strip(), fmt)
            return dt.strftime("%d-%b-%Y")
        except ValueError:
            pass
    return value


def extract_state_city(address_lines: list[str], zip_code: str) -> tuple[str, str]:
    """
    Extract US state abbreviation and city from multi-line address text.

    Strategy:
    1. Handle "CITY - X  STATE - XX" format (e.g. a "CITY - X  STATE - XX" style address)
    2. Work line-by-line from last line back, looking for a line that ends
       with a 2-letter state abbrev (optionally followed by ", USA" etc.)
    3. The city is the text before the state on that line, after stripping
       the street number/name tokens if the line also has a street address.
    """
    US_STATES = {
        "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR",
        "CALIFORNIA": "CA", "COLORADO": "CO", "CONNECTICUT": "CT",
        "DELAWARE": "DE", "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI",
        "IDAHO": "ID", "ILLINOIS": "IL", "INDIANA": "IN", "IOWA": "IA",
        "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA", "MAINE": "ME",
        "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI",
        "MINNESOTA": "MN", "MISSISSIPPI": "MS", "MISSOURI": "MO",
        "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
        "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM",
        "NEW YORK": "NY", "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND",
        "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
        "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC",
        "SOUTH DAKOTA": "SD", "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT",
        "VERMONT": "VT", "VIRGINIA": "VA", "WASHINGTON": "WA",
        "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
        "DISTRICT OF COLUMBIA": "DC",
    }
    STATE_ABBREVS = set(US_STATES.values())

    lines = [a.upper().strip() for a in address_lines if a and a.strip()]
    if not lines:
        return "", ""

    combined = " ".join(lines)

    # ── 0. Last line is a bare 2-letter state abbrev ──
    if len(lines) >= 2 and re.fullmatch(r"[A-Z]{2}", lines[-1]) and lines[-1] in STATE_ABBREVS:
        state = lines[-1]
        # City: last line before state, stripped of numbers/street terms
        city_line = lines[-2]
        # If city_line looks like a plain city name (no digits), use it directly
        if not re.search(r"\d", city_line):
            city = city_line.strip(",").strip()
        else:
            # Strip street address, grab last alpha word(s)
            tokens = city_line.split()
            city_tokens = []
            for tok in reversed(tokens):
                tok_clean = tok.rstrip(",")
                if re.match(r"^[A-Z]+$", tok_clean):
                    city_tokens.insert(0, tok_clean)
                else:
                    break
            city = " ".join(city_tokens) if city_tokens else city_line
        return state, city

    # ── 1. "CITY - X  STATE - XX" pattern ──
    sm = re.search(r"STATE\s*[-–]\s*([A-Z]{2,})", combined)
    if sm:
        token = sm.group(1).strip()
        state = US_STATES.get(token, token if token in STATE_ABBREVS else token)
        cm = re.search(r"CITY\s*[-–]\s*([A-Z][A-Z ]+?)(?:\s+STATE\b|\s*$)", combined)
        city = cm.group(1).strip().rstrip(",") if cm else ""
        return state, city

    # ── 2. Scan lines from last to first for one ending in a state abbrev ──
    for line in reversed(lines):
        # ── Check for full state name in line (e.g. "MERIDIAN, IDAHO, USA") ──
        line_no_country = re.sub(r",?\s*(USA|UNITED STATES|US)\s*$", "", line).strip()
        for full_name, abbrev in US_STATES.items():
            # Match full state name as a whole word/phrase at end or preceded by comma
            pattern = r"(.*?),\s*" + re.escape(full_name) + r"\s*$"
            fm = re.match(pattern, line_no_country)
            if fm:
                city_part = fm.group(1).strip().rstrip(",")
                if "," in city_part:
                    city_part = city_part.rsplit(",", 1)[-1].strip()
                return abbrev, city_part

        # Strip trailing junk
        line_clean = re.sub(r",?\s*(USA|UNITED STATES|US)\s*$", "", line).strip()

        # Try: "CITY, STATE" or "CITY STATE" at end of line
        # Find ALL 2-letter state matches on this line and take the last one
        all_matches = list(re.finditer(r"(?<![A-Z])([A-Z]{2})(?![A-Z])", line_clean))
        best_state = None
        best_city = None
        for m in reversed(all_matches):
            candidate_state = m.group(1)
            if candidate_state not in STATE_ABBREVS:
                continue
            # Skip common street direction abbreviations when they appear mid-address
            # only skip if there's more state-like content after them
            city_part = line_clean[:m.start()].strip().rstrip(",").strip()
            if not city_part:
                continue
            best_state = candidate_state
            best_city = city_part
            break

        if best_state:
            city = best_city
            if "," in city:
                city = city.rsplit(",", 1)[-1].strip()
                if city.upper() in US_STATES:
                    state_from_name = US_STATES[city.upper()]
                    city = best_city.rsplit(",", 1)[0].strip()
                    if "," in city:
                        city = city.rsplit(",", 1)[-1].strip()
                    return state_from_name, city
            # Remove leading street numbers
            if re.match(r"^\d+\s", city):
                tokens = city.split()
                city_tokens = []
                for tok in reversed(tokens):
                    tok_clean = tok.rstrip(",")
                    if re.match(r"^[A-Z]+$", tok_clean):
                        city_tokens.insert(0, tok_clean)
                    else:
                        break
                city = " ".join(city_tokens) if city_tokens else city
            return best_state, city

    # ── 3. Fallback: scan all tokens for a state abbrev ──
    tokens = re.split(r"[\s,]+", combined)
    for i, tok in enumerate(tokens):
        if tok in STATE_ABBREVS:
            # City: collect preceding alpha tokens until a digit or direction token
            city_tokens = []
            for t in reversed(tokens[:i]):
                t_clean = t.rstrip(",")
                if re.match(r"^[A-Z]+$", t_clean) and t_clean not in ("N", "S", "E", "W", "NE", "NW", "SE", "SW", "AVE", "ST", "DR", "RD", "LN", "BLVD", "PL", "WAY", "CT"):
                    city_tokens.insert(0, t_clean)
                else:
                    break
            city = " ".join(city_tokens)
            return tok, city

    return "", ""


# ---------------------------------------------------------------------------
# Document parser
# ---------------------------------------------------------------------------

def parse_application_form_document(docx_path: str) -> list[dict]:
    """
    Parse a structured Application Form .docx and return a list of applicant dicts.
    Each dict contains the columns needed for the output spreadsheet.
    """
    doc = Document(docx_path)
    records = []

    # We iterate over block-level elements (paragraphs + tables) in order.
    # Each applicant section ends with a paragraph containing the File No
    # (e.g. an alphanumeric reference code) followed by "View Application Form".

    current: dict = {}
    address_lines: list[str] = []     # collect multi-line address rows

    def flush_record():
        """Save the current record (if non-empty) and reset state."""
        nonlocal current, address_lines
        if current.get("name"):
            # Combine address lines → derive state & city
            state, city = extract_state_city(address_lines, current.get("zip", ""))
            current["state"] = state
            current["city"] = city
            records.append(current)
        current = {}
        address_lines = []

    # Walk all body elements in document order
    body = doc.element.body
    from docx.oxml.ns import qn

    elements = list(body)
    i = 0
    while i < len(elements):
        elem = elements[i]
        tag = elem.tag.split("}")[-1]  # 'p' or 'tbl'

        if tag == "p":
            from docx.text.paragraph import Paragraph
            para = Paragraph(elem, doc)
            text = clean(para.text)

            # Detect File No line: alphanumeric, length 12-15, no spaces
            if re.fullmatch(r"[A-Z0-9]{10,16}", text):
                # Assign file number to current record, then flush
                current["file_no"] = text
                flush_record()

        elif tag == "tbl":
            from docx.table import Table
            tbl = Table(elem, doc)
            # Each applicant table has specific rows; parse key-value pairs
            addr_row_count = 0
            for row in tbl.rows:
                cells = [clean(c.text) for c in row.cells]
                # The table has 4 columns: key1, val1, key2, val2
                # but some rows have merged cells
                if len(cells) < 2:
                    continue

                # Helper to get value at index safely
                def v(idx):
                    return cells[idx] if idx < len(cells) else ""

                k0 = v(0).lower()
                k2 = v(2).lower() if len(cells) > 2 else ""

                # ── Applicant name / Nationality ──
                if "applicant's name" in k0:
                    current["name"] = v(1)
                    if "nationality" in k2:
                        current["nationality"] = v(3)

                # ── Date of Birth / Birth Place ──
                elif "date of birth" in k0:
                    current["dob"] = parse_date(v(1))
                    if "birth place" in k2:
                        current["birth_place"] = v(3)

                # ── Sex / Phone ──
                elif k0 == "sex":
                    current["sex"] = v(1)

                # ── Occupation ──
                elif "occupation" in k0:
                    current["occupation"] = v(1)

                # ── Address (3 possible rows, first col empty for rows 2-3) ──
                elif "present address" in k0:
                    addr_row_count = 0
                    addr_line = v(1)
                    if addr_line:
                        address_lines.append(addr_line)
                    addr_row_count += 1

                elif k0 == "" and addr_row_count > 0 and addr_row_count < 4:
                    # Continuation row of address
                    addr_line = v(1)
                    if addr_line:
                        address_lines.append(addr_line)
                    addr_row_count += 1

                # ── Zip code ──
                elif "zip" in k0 or "pin" in k0 or "area code" in k0:
                    current["zip"] = v(1)

                # ── Mobile ──
                elif "mobile" in k0:
                    current["mobile"] = strip_phone(v(1))
                    if "email" in k2:
                        current["email"] = v(3)

                # ── Amount / Currency ──
                elif k0 == "amount":
                    current["amount"] = v(1)
                    if "currency" in k2:
                        current["currency"] = v(3)

        i += 1

    # Flush any trailing record that didn't have a File No paragraph after it
    if current.get("name"):
        state, city = extract_state_city(address_lines, current.get("zip", ""))
        current["state"] = state
        current["city"] = city
        records.append(current)

    return records


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

COLUMNS = [
    ("File No",        "file_no"),
    ("Applicant's Name", "name"),
    ("Date of Birth",  "dob"),
    ("Gender",         "sex"),
    ("Mobile/Tel No",  "mobile"),
    ("E-Mail",         "email"),
    ("State",          "state"),
    ("City",           "city"),
    ("Occupation",     "occupation"),
    ("Employer Name",  "employer"),   # not present in source; left blank
]


def write_excel(records: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Application Extraction"

    # ── Header styling ──
    header_fill = PatternFill("solid", fgColor="1F4E79")   # dark blue
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [18, 28, 14, 8, 14, 32, 7, 16, 18, 18]

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for col_idx, (cell, width) in enumerate(zip(ws[1], col_widths), start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30

    # ── Data rows ──
    data_font = Font(size=10)
    data_align = Alignment(vertical="center")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")   # light blue

    for row_idx, rec in enumerate(records, start=2):
        row_data = [rec.get(field, "") for _, field in COLUMNS]
        ws.append(row_data)
        for cell in ws[row_idx]:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"✅  Saved {len(records)} records → {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python docx_to_excel_extractor.py <input.docx> [output.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        stem = Path(input_path).stem
        output_path = str(Path(input_path).parent / f"{stem}_extracted.xlsx")

    print(f"📄  Parsing: {input_path}")
    records = parse_application_form_document(input_path)
    print(f"🔍  Found {len(records)} applicant records")
    write_excel(records, output_path)


if __name__ == "__main__":
    main()
